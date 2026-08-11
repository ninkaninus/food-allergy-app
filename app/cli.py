"""
Kommandolinje. Kør inde i containeren:

    docker compose exec allergiscan python -m app.cli adduser dig@example.dk "William"
    docker compose exec allergiscan python -m app.cli reindex
    docker compose run --rm allergiscan python -m app.cli migrate /data/allergiscan.db
"""

from __future__ import annotations

import asyncio
import getpass
import os
import re
import sys
import tempfile

import httpx

from sqlalchemy import create_engine, func, select, text

from . import ingredients as ix
from .auth import hash_password, validate_new_password
from .db import DATABASE_URL, SessionLocal, default_household, init_db
from .models import Base, ImportedProduct, Product, User


def adduser(email: str, name: str, role: str = "admin") -> None:
    pw = getpass.getpass("Adgangskode (mindst 14 tegn): ")
    if pw != getpass.getpass("Gentag: "):
        sys.exit("De to kodeord er ikke ens.")
    try:
        asyncio.run(validate_new_password(pw))
    except Exception as e:
        sys.exit(getattr(e, "detail", str(e)))

    init_db()
    with SessionLocal() as db:
        if db.scalar(select(User).where(User.email == email.lower())):
            sys.exit("Den mail findes allerede.")
        db.add(
            User(
                household_id=default_household(db).id,
                email=email.lower(),
                name=name,
                password_hash=hash_password(pw),
                role=role,
                source="local",
            )
        )
        db.commit()
    print(f"Oprettet {email} som {role}.")


def reindex() -> None:
    """Genopbygger ingrediensindekset fra de rå tekster, vi allerede har."""
    init_db()
    with SessionLocal() as db:
        n = 0
        for p in db.scalars(select(Product)):
            n += bool(ix.index_product(db, p.ean, None, p.off_allergen_tags,
                                       p.ingredients_text))
        db.commit()
    print(f"Genindekserede {n} varer.")


def _eksport_url(kilde: str) -> str:
    """Et Google Sheets-link oversættes til dets xlsx-eksport; alle andre
    URL'er bruges som de er."""
    m = re.match(r"https?://docs\.google\.com/spreadsheets/d/([\w-]+)", kilde)
    if m:
        return f"https://docs.google.com/spreadsheets/d/{m.group(1)}/export?format=xlsx"
    return kilde


def _hent_liste(url: str) -> str:
    """Henter regnearket til en midlertidig fil og returnerer stien.
    Kræver at arket er delt med 'alle med linket kan se'."""
    r = httpx.get(_eksport_url(url), follow_redirects=True, timeout=120)
    r.raise_for_status()
    if not r.content.startswith(b"PK"):  # xlsx er et zip-arkiv
        sys.exit(
            "Det hentede er ikke en xlsx-fil — er arket delt med "
            "'alle med linket kan se'? (Google viser ellers en loginside.)"
        )
    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    f.write(r.content)
    f.close()
    return f.name


VALIDERET_MOD = "æg, mælk, tomat og banan"


def import_liste(kilde: str | None = None, valideret_mod: str | None = None) -> None:
    """
    Importerer jeres gamle godkendt-liste fra regnearket (xlsx).

    Kilden kan være en lokal fil ELLER en URL — et Google Sheets-link
    hentes selv (kræver 'alle med linket kan se'). Uden argument bruges
    LISTE_URL fra miljøet (.env), så genimport er én kommando.

    ALLE varer på listen var valideret, før de kom ind i regnearket —
    uden æg, mælk, tomat og banan. Arkets "Valideret"-kolonne er et dødt
    felt og ignoreres. Betydningen kan overstyres med andet argument.

    Arket har ingen EAN-koder, så rækkerne bliver IKKE til domme — de
    lander i en separat opslagsliste, som appen søger i og viser som
    hint ved scanning. Hver vare bliver først grøn, når den scannes og
    bekræftes mod den fysiske emballage, som alle andre.

    Genimport udskifter hele listen — tabellen ejes af importen og
    genskabes hver gang, så også skemaændringer heler sig selv.
    """
    from openpyxl import load_workbook

    kilde = kilde or os.getenv("LISTE_URL")
    if not kilde:
        sys.exit(
            "Angiv en fil eller URL — eller sæt LISTE_URL i .env til "
            "regnearkets Google Sheets-link, så husker kommandoen den selv."
        )
    valideret_mod = valideret_mod or VALIDERET_MOD
    fil = _hent_liste(kilde) if "://" in kilde else kilde

    init_db()
    wb = load_workbook(fil, read_only=True, data_only=True)
    rows: list[dict] = []
    for ws in wb:
        if ws.title.strip().lower() == "masterdata":
            continue
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            continue
        kol = {str(h).strip().lower(): i for i, h in enumerate(header) if h}
        if "beskrivelse" not in kol:
            continue

        def felt(row, navn):
            i = kol.get(navn)
            if i is None or i >= len(row) or row[i] is None:
                return None
            s = str(row[i]).strip()
            return s or None

        for row in it:
            navn = felt(row, "beskrivelse")
            if not navn:
                continue
            rows.append(dict(
                kategori=ws.title.strip(),
                navn=navn,
                producent=felt(row, "producent"),
                butik=felt(row, "butik"),
                link=felt(row, "link"),
                erstatning_for=felt(row, "bruges  fx som erstatning for")
                               or felt(row, "bruges fx som erstatning for"),
                valideret=True,
                valideret_mod=valideret_mod,
            ))

    # Tabellen ejes af importen og indeholder kun importerede rækker —
    # genskab den, så nye kolonner ikke kræver håndholdt migrering.
    from .db import engine
    tabel = ImportedProduct.__table__
    tabel.drop(engine, checkfirst=True)
    tabel.create(engine)

    with SessionLocal() as db:
        hh = default_household(db)
        for r in rows:
            db.add(ImportedProduct(household_id=hh.id, **r))
        db.commit()
    print(f"Importerede {len(rows)} varer — alle bekræftet uden {valideret_mod}.")


def migrate(kilde: str, maal: str | None = None) -> None:
    """
    Kopierer ALLE tabeller fra én database til en anden — typisk fra
    SQLite-filen til Postgres-containeren. Målet er den konfigurerede
    DATABASE_URL, medmindre andet angives, og det SKAL være tomt: har
    appen allerede startet én gang mod målet, har init_db seedet det,
    og så stopper vi hellere end at duplikere. Se deploy/UNRAID.md.
    """
    if "://" not in kilde:
        kilde = f"sqlite:///{kilde}"
    maal = maal or DATABASE_URL
    if kilde == maal:
        sys.exit("Kilde og mål er den samme database — ingenting at gøre.")

    k_eng = create_engine(kilde)
    m_eng = create_engine(maal)
    Base.metadata.create_all(m_eng)

    with m_eng.connect() as m:
        for t in Base.metadata.sorted_tables:
            if m.execute(select(func.count()).select_from(t)).scalar():
                sys.exit(
                    f"Målet er ikke tomt ({t.name} har rækker). Migrér kun til en "
                    "frisk database — se fremgangsmåden i deploy/UNRAID.md."
                )

    # Forkontrol af fremmednøgler. SQLite håndhæver dem ikke; Postgres gør.
    # En database, der har virket i månedsvis, kan derfor have brudte
    # referencer — og dem vil vi vide om FØR flytningen, ikke som en
    # halvvejs mislykket indsættelse med et stakspor.
    brud = []
    with k_eng.connect() as k:
        for t in Base.metadata.sorted_tables:
            for fk in t.foreign_keys:
                barn, foraelder = fk.parent, fk.column
                mangler = k.execute(
                    select(barn, func.count())
                    .where(barn.isnot(None))
                    .where(barn.notin_(select(foraelder)))
                    .group_by(barn)
                ).all()
                if mangler:
                    brud.append((t.name, barn.name, foraelder.table.name, mangler))
    if brud:
        print("Kildedatabasen har referencer, der peger på rækker, som ikke findes:")
        for tabel, kol, mod, mangler in brud:
            vaerdier = ", ".join(f"{v} ({n}x)" for v, n in mangler[:5])
            mere = f" og {len(mangler) - 5} flere" if len(mangler) > 5 else ""
            print(f"  {tabel}.{kol} -> {mod}: {vaerdier}{mere}")
        sys.exit(
            "Flytningen er ikke påbegyndt — intet er ændret. Ret referencerne "
            "i kilden, eller opdatér appen, hvis begrænsningen er forkert."
        )

    total = 0
    with k_eng.connect() as k, m_eng.begin() as m:
        for t in Base.metadata.sorted_tables:
            rows = [dict(r._mapping) for r in k.execute(t.select())]
            if rows:
                m.execute(t.insert(), rows)
            print(f"  {t.name}: {len(rows)} rækker")
            total += len(rows)

    # Postgres tildeler id'er fra sekvenser, og de står stadig på nul
    # efter indsættelse med eksplicitte id'er — stil dem frem.
    if m_eng.dialect.name == "postgresql":
        with m_eng.begin() as m:
            for t in Base.metadata.sorted_tables:
                pk = list(t.primary_key.columns)
                if len(pk) == 1 and pk[0].name == "id":
                    m.execute(text(
                        f"SELECT setval(pg_get_serial_sequence('{t.name}', 'id'), "
                        f"COALESCE((SELECT MAX(id) FROM {t.name}), 1))"
                    ))

    print(f"Færdig: {total} rækker flyttet til {m_eng.url.render_as_string(hide_password=True)}")


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else ""
    if cmd == "adduser":
        adduser(*sys.argv[2:])
    elif cmd == "reindex":
        reindex()
    elif cmd == "migrate":
        migrate(*sys.argv[2:])
    elif cmd in ("import", "import-liste"):
        import_liste(*sys.argv[2:])
    else:
        sys.exit(
            "Brug: python -m app.cli [adduser EMAIL NAVN [ROLLE] | reindex "
            "| migrate KILDE [MÅL] | import [FIL.xlsx eller URL]]"
        )
