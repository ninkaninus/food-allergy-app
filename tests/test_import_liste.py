"""
Import af den gamle godkendt-liste (regneark uden EAN).

Det vigtigste, testene skal holde fast i: importen laver OPSLAGSRÆKKER,
aldrig domme. Verdict-tabellen skal være urørt efter import — listen
kender ingen stregkoder, og grøn kræver stadig et menneske med pakken
i hånden.
"""
import os
import pathlib
import sys
import tempfile

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))

os.environ.setdefault("DATA_DIR", tempfile.mkdtemp())
os.environ.setdefault(
    "RULES_PATH",
    str(pathlib.Path(__file__).resolve().parents[1] / "data" / "allergens.yaml"),
)
os.environ.setdefault("COOKIE_SECURE", "0")

from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import func, select

from app.cli import import_liste
from app.db import SessionLocal, default_household, init_db
from app.main import _gammel_liste_hint, app
from app.models import ImportedProduct, Product, Verdict, now


def _ark(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Kiks P"
    ws.append(["Beskrivelse", "Producent", "Billede", "Butik", "Link", "I valgt butik", "Valideret"])
    ws.append(["Digestive", "Rema 1000", None, "Rema1000", "https://x", "Ja", "x"])
    ws.append(["Marie kiks", "Coop", None, "Super Brugsen", None, "Ja", None])
    ws.append([None, None, None, None, None, None, None])   # tom række springes over

    e = wb.create_sheet("Erstatningsprodukter")
    e.append(["Beskrivelse", "Producent", "Billede", "Butik", "Bruges  fx som erstatning for", "Link", "I valgt butik", "Valideret"])
    e.append(["Block (Smør)", "Naturlig", None, "Rema1000", "Smør", "https://y", "Ja", "x"])

    m = wb.create_sheet("MasterData")   # skal ignoreres
    m.append(["Butikker"])
    m.append(["Netto"])

    sti = tmp_path / "liste.xlsx"
    wb.save(sti)
    return str(sti)


def test_import_og_genimport(tmp_path):
    init_db()
    fil = _ark(tmp_path)
    import_liste(fil)
    import_liste(fil)   # genimport må ikke duplikere

    with SessionLocal() as db:
        hh = default_household(db)
        rows = list(db.scalars(
            select(ImportedProduct).where(ImportedProduct.household_id == hh.id)
        ))
        assert len(rows) == 3
        digestive = next(r for r in rows if r.navn == "Digestive")
        assert digestive.producent == "Rema 1000"
        assert digestive.kategori == "Kiks P"
        # Arkets "Valideret"-kolonne er et dødt felt: ALT på listen var
        # valideret (uden æg, mælk, tomat og banan), før det kom ind.
        assert all(r.valideret for r in rows)
        assert all(r.valideret_mod == "æg, mælk, tomat og banan" for r in rows)
        smoer = next(r for r in rows if r.navn == "Block (Smør)")
        assert smoer.erstatning_for == "Smør"
        assert not any(r.kategori == "MasterData" for r in rows)


def test_butik_kolonnen_ignoreres(tmp_path):
    """Arket har stadig en Butik-kolonne, men butik er ikke data, appen får
    om fremtidige varer: en scannet vare har ingen butik. Et filter, der
    kun kender arkets rækker, ville skjule alt det, I har scannet."""
    init_db()
    import_liste(_ark(tmp_path))
    with SessionLocal() as db:
        rows = list(db.scalars(select(ImportedProduct)))
    assert rows and not any(hasattr(r, "butik") for r in rows)


def test_import_skaber_ingen_domme(tmp_path):
    """Invariantens naboklausul: listen har ingen EAN og må aldrig blive
    til domme. Grøn kræver stadig et menneske."""
    init_db()
    with SessionLocal() as db:
        foer = db.scalar(select(func.count()).select_from(Verdict))
    import_liste(_ark(tmp_path))
    with SessionLocal() as db:
        efter = db.scalar(select(func.count()).select_from(Verdict))
    assert efter == foer


def test_google_link_oversaettes_til_eksport_url():
    from app.cli import _eksport_url
    url = "https://docs.google.com/spreadsheets/d/1NrjeAbC_-xyz/edit?usp=sharing"
    assert _eksport_url(url) == (
        "https://docs.google.com/spreadsheets/d/1NrjeAbC_-xyz/export?format=xlsx"
    )
    # andre URL'er og lokale stier røres ikke
    assert _eksport_url("https://x.dk/liste.xlsx") == "https://x.dk/liste.xlsx"


def test_import_fra_url_henter_selv(tmp_path, monkeypatch):
    """URL-kilder hentes med httpx; en HTML-loginside afvises tydeligt."""
    import app.cli as cli

    xlsx = open(_ark(tmp_path), "rb").read()

    class Svar:
        def __init__(self, content): self.content = content
        def raise_for_status(self): pass

    monkeypatch.setattr(cli.httpx, "get", lambda *a, **k: Svar(xlsx))
    init_db()
    cli.import_liste("https://docs.google.com/spreadsheets/d/abc123/edit")
    with SessionLocal() as db:
        hh = default_household(db)
        antal = db.scalar(
            select(func.count()).select_from(ImportedProduct)
            .where(ImportedProduct.household_id == hh.id)
        )
    assert antal == 3

    import pytest as _pytest
    monkeypatch.setattr(cli.httpx, "get", lambda *a, **k: Svar(b"<html>login"))
    with _pytest.raises(SystemExit):
        cli.import_liste("https://docs.google.com/spreadsheets/d/abc123/edit")


def test_hint_matcher_paa_navn_og_producent(tmp_path):
    init_db()
    import_liste(_ark(tmp_path))
    with SessionLocal() as db:
        hh = default_household(db)
        # to betydende ord fælles ("marie" + "kiks")
        hits = _gammel_liste_hint(db, hh.id, "Marie kiks med fuldkorn", "Coop")
        assert any(h["navn"] == "Marie kiks" for h in hits)
        # ét ord + producentmatch
        hits = _gammel_liste_hint(db, hh.id, "Digestive", "Rema 1000")
        assert any(h["navn"] == "Digestive" for h in hits)
        # ingenting til fælles
        assert _gammel_liste_hint(db, hh.id, "Piskefløde", "Arla") == []


def test_scan_hintet_bruger_kun_offs_navn_ikke_familiens_eget(tmp_path):
    """
    `/api/scan` sendte en overgang `_visningsnavn(product)` (som lader
    familiens eget `navn_manuelt` vinde) videre til _gammel_liste_hint().
    Hintet er BEROLIGENDE ("bekræftet uden … — men tjek stadig
    emballagen"), og den farlige retning er under-advarsel: et navn,
    familien selv har fundet på til én vare, kan tilfældigt dele to ord
    med en HELT ANDEN vare på det gamle ark. Kun Open Food Facts' eget,
    ikke-familiestyrede navn må udløse hintet.
    """
    init_db()
    import_liste(_ark(tmp_path))   # rækken "Marie kiks" / "Coop"

    # product.name (OFF's eget, urelateret) matcher IKKE noget på arket —
    # navn_manuelt (familiens eget) deler derimod to ord ("marie"+"kiks")
    # med "Marie kiks"/Coop. Ingen brand, så der ikke sniger sig et
    # brand-alene-match ind (se test_hint_matcher_paa_navn_og_producent).
    ean = "5701234599901"
    with SessionLocal() as db:
        db.add(Product(
            ean=ean, name="Hapithun i olie", navn_manuelt="Marie kiks med fuldkorn",
            brand=None, ingredients_text="Hvedemel, vand, salt.", source="off",
            fetched_at=now().replace(tzinfo=None),   # undgår et rigtigt OFF-kald
        ))
        db.commit()

    c = TestClient(app)
    r = c.get(f"/api/scan/{ean}")
    assert r.status_code == 200, r.text
    d = r.json()
    assert d["product"]["name"] == "Marie kiks med fuldkorn", (
        "visningsnavnet (product.name i svaret) skal stadig vise familiens eget navn"
    )
    assert d["gammel_liste"] == [], (
        "hintet reagerede på familiens eget navn (navn_manuelt) — det må "
        "kun Open Food Facts' eget navn kunne"
    )


def test_genimport_bevarer_stregkode_koblinger(tmp_path):
    """
    Var i drift: import_liste droppede tabellen og indsatte rækker uden
    ean, så hver genimport slettede de stregkoder, et menneske havde
    knyttet — og ROADMAP'en opfordrer direkte til at genimportere, når
    arket opdateres. Koblingen er familiens eget arbejde, ikke arkets.
    """
    init_db()
    fil = _ark(tmp_path)
    import_liste(fil)

    with SessionLocal() as db:
        row = db.scalar(select(ImportedProduct).where(ImportedProduct.navn == "Digestive"))
        row.ean = "5700000000031"
        db.commit()

    import_liste(fil)

    with SessionLocal() as db:
        row = db.scalar(select(ImportedProduct).where(ImportedProduct.navn == "Digestive"))
        assert row.ean == "5700000000031", "genimport slettede koblingen"
        # De øvrige rækker må ikke arve en stregkode, de aldrig fik.
        andre = db.scalars(
            select(ImportedProduct).where(ImportedProduct.navn != "Digestive")
        ).all()
        assert all(r.ean is None for r in andre)


def test_to_raekker_med_samme_noegle_afbryder_foer_droppet(tmp_path):
    """
    Nøglen (navn, producent, kategori) er ikke unik. To rækker med samme
    nøgle ville begge få den sidste stregkode — en stille sammenblanding
    af to varer. Importen SLETTER tabellen, så det skal fanges før.
    """
    import pytest

    init_db()
    fil = _ark(tmp_path)
    import_liste(fil)

    with SessionLocal() as db:
        raekker = db.scalars(
            select(ImportedProduct).where(ImportedProduct.navn == "Digestive")
        ).all()
        # Lav en dublet af nøglen med en ANDEN stregkode.
        first = raekker[0]
        db.add(ImportedProduct(
            household_id=first.household_id, navn=first.navn,
            producent=first.producent, kategori=first.kategori,
            ean="5700000000055",
        ))
        first.ean = "5700000000062"
        db.commit()
        foer = db.query(ImportedProduct).count()

    with pytest.raises(SystemExit) as e:
        import_liste(fil)
    assert "deler nøglen" in str(e.value)

    with SessionLocal() as db:
        assert db.query(ImportedProduct).count() == foer, \
            "tabellen blev droppet, selvom importen afbrød"
